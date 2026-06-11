"""
Inspect APR baza-podataka page and check Vracena Posta XLS for emails.
"""
import io, re, ssl, sys, urllib.request
from bs4 import BeautifulSoup

sys.stdout.reconfigure(encoding="utf-8")
ctx = ssl._create_unverified_context()
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept-Language": "sr-RS,sr;q=0.9",
    "Accept": "text/html,*/*",
}


def get(url):
    req = urllib.request.Request(url, headers=HEADERS)
    try:
        with urllib.request.urlopen(req, context=ctx, timeout=20) as resp:
            raw = resp.read()
            print(f"  GET {url[:80]} -> {resp.status}  len={len(raw)}")
            return raw
    except Exception as e:
        print(f"  GET {url[:80]} -> ERROR: {e}")
        return b""


# 1) Parse baza-podataka page fully
print("=== baza-podataka page ===")
html = get("https://www.apr.gov.rs/registri/privredni-subjekti/baza-podataka.1275.html")
if html:
    soup = BeautifulSoup(html, "lxml")
    # All links with file extensions or "download"
    for a in soup.find_all("a", href=True):
        href = a["href"]
        text = a.get_text(strip=True)[:60]
        if any(ext in href.lower() for ext in [".xls",".xlsx",".csv",".zip",".xml",".json","download","baza","export"]):
            print(f"  [{text}] -> {href}")
    # All text content on the page (look for database descriptions)
    paragraphs = soup.find_all(["p","h1","h2","h3","h4","li"])
    for p in paragraphs[:30]:
        t = p.get_text(strip=True)
        if t and len(t) > 20:
            print(f"  TEXT: {t[:120]}")

# 2) Try to download the Vracena Posta XLS (it's a list of businesses with returned mail)
print("\n=== Vracena Posta XLS (returned mail list) ===")
xls_data = get("https://pretraga5.apr.gov.rs/vracenaposta/Vracena%20Posta.xls")
if xls_data and len(xls_data) > 1000:
    print(f"  Downloaded {len(xls_data)} bytes")
    # Try to read with openpyxl/xlrd
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(xls_data))
        ws = wb.active
        print(f"  Columns: {[ws.cell(1,c).value for c in range(1, ws.max_column+1)]}")
        print(f"  Sample rows:")
        for row in ws.iter_rows(min_row=2, max_row=5, values_only=True):
            print(f"    {row}")
    except Exception as e1:
        print(f"  openpyxl failed: {e1}")
        try:
            import xlrd
            wb = xlrd.open_workbook(file_contents=xls_data)
            ws = wb.sheet_by_index(0)
            print(f"  Columns: {ws.row_values(0)}")
            for i in range(1, min(5, ws.nrows)):
                print(f"  Row {i}: {ws.row_values(i)}")
        except Exception as e2:
            print(f"  xlrd also failed: {e2}")
            # Dump raw text
            text = xls_data.decode("cp1251", errors="replace")
            emails = re.findall(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}", text)
            print(f"  Emails in XLS: {emails[:20]}")

# 3) Check the BD database download page more carefully
print("\n=== Check for actual BD database downloads ===")
bd_html = get("https://www.apr.gov.rs/registri/privredni-subjekti/baza-podataka.1275.html")
if bd_html:
    # Decode and search for download URLs
    text = bd_html.decode("utf-8", errors="replace")
    all_links = re.findall(r'href=["\']([^"\']{10,200})["\']', text)
    interesting = [l for l in all_links if any(
        k in l.lower() for k in ["baza","download","export","xls","csv","xml","zip","pretraga5","pretraga2","pretraga3"]
    )]
    print(f"  Interesting links: {interesting[:20]}")
